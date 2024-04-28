import openpyxl

wb = openpyxl.load_workbook('./static/files/weather_data.xlsx')

sheet = wb.active

iter_rows = sheet.iter_rows()

for item in iter_rows:
    print(item)

