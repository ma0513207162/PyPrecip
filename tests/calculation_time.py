from typing import List 
from openpyxl import load_workbook


def test_unit_func():
    reader_list:List[List[str]] = []

    workbook = load_workbook(filename = "../../static/test_data.xlsx", data_only=True)
    sheet = workbook["Sheet1"]

    for row in sheet.iter_rows(values_only=True):
        reader_list.append(list(row))

    return reader_list; 

def test_unit_func2():
    reader_list:List[List[str]] = []

    workbook = load_workbook(filename = "../../static/test_data.xlsx", data_only=True)
    sheet = workbook["Sheet1"]

    for row in sheet.iter_rows(values_only=True):
        row_data = [] 
        for cell in row:
            row_data.append(cell)
        reader_list.append(row_data)
    return reader_list; 











