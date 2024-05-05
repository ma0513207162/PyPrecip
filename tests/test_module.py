import openpyxl 

# 加载 Excel 文件 
workbook = openpyxl.load_workbook("./static/test_data.xlsx"); 

# 获取工作表 
worksheet = workbook.active; 

sheet_iter_rows: list = list(worksheet.iter_rows(values_only = True)) 

sheet_iter_rows = list(zip(*sheet_iter_rows))

print(sheet_iter_rows)
